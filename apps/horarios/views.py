from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
import json
import logging
from .models import Horario, BloqueHorario, Aula
from inscripciones.models import Asignatura, PeriodoAcademico
from usuarios.models import Usuario
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

logger = logging.getLogger(__name__)

# Cargos excluidos de la nómina docente
_CARGOS_EXCLUIDOS_NOMINA = ['Est. Limpieza', 'Mantenimiento']

def dashboard_horarios(request, horario_id=None):
    periodos = PeriodoAcademico.objects.all()
    asignaturas_qs = Asignatura.objects.all().order_by('ano_grado', 'nombre')
    docentes = Usuario.objects.filter(rol='DOCENTE')  # FK legacy (no mostrar en UI)

    # Serializar asignaturas como JSON para filtrado dinámico en el frontend
    asignaturas_json = json.dumps([
        {'id': a.id, 'codigo': a.codigo, 'nombre': a.nombre, 'ano_grado': a.ano_grado}
        for a in asignaturas_qs
    ], ensure_ascii=False)

    # Personal de nómina válido para campo Docente (excluye Limpieza y Mantenimiento)
    try:
        from pagos.models import Personal
        docentes_nomina = Personal.objects.filter(activo=True).exclude(
            cargo__in=_CARGOS_EXCLUIDOS_NOMINA
        ).order_by('cargo', 'nombre_completo')
    except Exception:
        docentes_nomina = []

    # Serializar todos los docentes disponibles como JSON para el frontend
    # Combina Personal (nómina) y docentes de Usuario como fallback
    docentes_list = []
    seen_names = set()
    for p in docentes_nomina:
        if p.nombre_completo not in seen_names:
            docentes_list.append({
                'id': p.id,
                'nombre': p.nombre_completo,
                'cargo': p.cargo,
                'source': 'nomina',
            })
            seen_names.add(p.nombre_completo)
    for d in docentes:
        nombre = d.nombre_completo or f"{d.first_name} {d.last_name}".strip() or d.username
        if nombre not in seen_names:
            docentes_list.append({
                'id': d.id,
                'nombre': nombre,
                'cargo': 'Docente',
                'source': 'usuario',
            })
            seen_names.add(nombre)
    docentes_json = json.dumps(docentes_list, ensure_ascii=False)

    modo_lectura = False
    horario_precargado = None

    if horario_id:
        horario_precargado = get_object_or_404(Horario, id=horario_id)
        if request.GET.get('edit') != '1':
            modo_lectura = True

    # Horarios existentes para acceso rápido en la barra lateral
    horarios_existentes = Horario.objects.select_related('periodo').order_by(
        'ano_grado', 'seccion'
    )

    context = {
        'periodos': periodos,
        'asignaturas': asignaturas_qs,
        'asignaturas_json': asignaturas_json,
        'docentes': docentes,
        'docentes_nomina': docentes_nomina,
        'docentes_json': docentes_json,
        'modo_lectura': modo_lectura,
        'horario_precargado': horario_precargado,
        'horarios_existentes': horarios_existentes,
    }
    return render(request, 'horarios/dashboard.html', context)

def lista_horarios(request):
    # Agrupamos los horarios por año
    horarios = Horario.objects.all().order_by('ano_grado', 'seccion')
    agrupados = {
        1: [], 2: [], 3: [], 4: [], 5: []
    }
    for h in horarios:
        if h.ano_grado in agrupados:
            agrupados[h.ano_grado].append(h)
        else:
            agrupados[h.ano_grado] = [h]
            
    context = {
        'agrupados': agrupados
    }
    return render(request, 'horarios/lista.html', context)

def api_obtener_horarios(request):
    """Devuelve los turnos de un horario especifico para FullCalendar"""
    horario_id = request.GET.get('horario_id')
    if not horario_id:
        return JsonResponse([], safe=False)

    bloques = BloqueHorario.objects.filter(horario_id=horario_id)
    eventos = []
    
    for b in bloques:
        title = ""
        if b.tipo == 'CLASE':
            asignatura_nombre = b.asignatura.nombre if b.asignatura else 'Sin Materia'
            # Priorizar docente_nombre (nómina) sobre FK legacy
            docente_nombre = b.docente_nombre or (
                f"{b.docente.first_name} {b.docente.last_name}" if b.docente else 'Sin Docente'
            )
            # Priorizar aula_numero sobre FK legacy
            aula_label = f"Aula {b.aula_numero}" if b.aula_numero else (
                b.aula.nombre if b.aula else ''
            )
            title = f"{asignatura_nombre}\n{docente_nombre}\n{aula_label}"
        else:
            title = b.get_tipo_display()

        eventos.append({
            'id': b.id,
            'title': title,
            'daysOfWeek': [int(b.dia_semana)],
            'startTime': b.hora_inicio.strftime('%H:%M:%S'),
            'endTime': b.hora_fin.strftime('%H:%M:%S'),
            'color': b.color_hex,
            'extendedProps': {
                'tipo': b.tipo,
                'asignatura_id': b.asignatura.id if b.asignatura else None,
                'docente_nombre': b.docente_nombre or '',
                'aula_numero': b.aula_numero or '',
                # Legacy fallback
                'docente_id': b.docente.id if b.docente else None,
                'aula_id': b.aula.id if b.aula else None,
            }
        })
        
    return JsonResponse(eventos, safe=False)

@csrf_exempt
def api_guardar_bloque(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            periodo_id = data.get('periodo_id')
            ano = data.get('ano')
            seccion = data.get('seccion')
            
            horario, created = Horario.objects.get_or_create(
                periodo_id=periodo_id,
                ano_grado=ano,
                seccion=seccion
            )

            bloque_id = data.get('bloque_id')
            tipo = data.get('tipo', 'CLASE')
            dia = data.get('dia')
            hora_inicio = data.get('hora_inicio')
            hora_fin = data.get('hora_fin')
            color = data.get('color', '#6366F1')
            
            if bloque_id:
                bloque = BloqueHorario.objects.get(id=bloque_id)
            else:
                bloque = BloqueHorario(horario=horario)
            
            bloque.tipo = tipo
            bloque.dia_semana = dia
            bloque.hora_inicio = hora_inicio
            bloque.hora_fin = hora_fin
            bloque.color_hex = color

            # Validar solapamiento de bloques en el mismo horario
            overlap = BloqueHorario.objects.filter(
                horario=horario,
                dia_semana=dia,
                hora_inicio__lt=hora_fin,
                hora_fin__gt=hora_inicio
            )
            if bloque_id:
                overlap = overlap.exclude(id=bloque_id)
                
            if overlap.exists():
                return JsonResponse({'status': 'error', 'message': 'Las horas seleccionadas ya están ocupadas por otro bloque en este horario.'}, status=400)

            if tipo == 'CLASE':
                if data.get('asignatura_id'):
                    bloque.asignatura_id = data.get('asignatura_id')
                else:
                    bloque.asignatura = None

                # Guardar docente desde nómina (texto) — campo prioritario
                bloque.docente_nombre = str(data.get('docente_nombre', '')).strip()
                if data.get('docente_id'):
                    bloque.docente_id = data.get('docente_id')
                else:
                    bloque.docente = None


                # Guardar aula como número (solo dígitos) — validación backend
                raw_aula = str(data.get('aula_numero', '')).strip()
                if raw_aula and not raw_aula.isdigit():
                    return JsonResponse({'status': 'error',
                                         'message': 'El número de aula debe contener solo dígitos.'}, status=400)
                bloque.aula_numero = raw_aula
            else:
                bloque.asignatura    = None
                bloque.docente       = None
                bloque.docente_nombre = ''
                bloque.aula          = None
                bloque.aula_numero   = ''

            bloque.save()
            logger.info(
                f"[Horarios] Bloque guardado id={bloque.id} tipo={tipo} "
                f"docente='{bloque.docente_nombre}' aula='{bloque.aula_numero}'"
            )
            
            # Auditoria
            from auditoria.models import registrar_evento
            accion = 'MODIFICACION' if bloque_id else 'CREACION'
            registrar_evento(accion, f"Bloque de horario {'modificado' if bloque_id else 'creado'}: {tipo} ({horario.ano_grado}-{horario.seccion})", 'Horarios', request.user.username, 'INFORMATIVO')
            
            return JsonResponse({'status': 'ok', 'bloque_id': bloque.id, 'horario_id': horario.id})
        except Exception as e:
            logger.error(f"[Horarios] Error guardando bloque: {e}", exc_info=True)
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

@csrf_exempt
def api_eliminar_bloque(request, bloque_id):
    from auditoria.models import registrar_evento
    
    if request.method == 'POST':
        try:
            b = BloqueHorario.objects.get(id=bloque_id)
            registrar_evento('INACTIVACION', f"Bloque de horario eliminado: {b.tipo} ({b.horario.ano_grado}-{b.horario.seccion})", 'Horarios', request.user.username, 'MEDIO')
            b.delete()
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            logger.error(f"[Horarios] Error al eliminar bloque: {e}", exc_info=True)
            pass
    return JsonResponse({'status': 'error'})

def exportar_excel(request, horario_id):
    horario = get_object_or_404(Horario, id=horario_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Horario {horario.ano_grado}-{horario.seccion}"

    # Encabezado
    ws.merge_cells('A1:F1')
    ws['A1'] = f"HORARIO DE CLASES: {horario.ano_grado}° '{horario.seccion}' - PERÍODO: {horario.periodo.nombre}"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    dias = ['Hora', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']
    for col, dia in enumerate(dias, start=1):
        cell = ws.cell(row=2, column=col, value=dia)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Recopilar datos
    bloques = horario.bloques.all().order_by('hora_inicio')
    
    # Set ancho columnas
    ws.column_dimensions['A'].width = 15
    for c in "BCDEF":
        ws.column_dimensions[c].width = 25
        
    # Agrupar por hora de inicio única
    horas = sorted(list(set(b.hora_inicio for b in bloques)))
    
    row_idx = 3
    for h_inicio in horas:
        h_formatted = h_inicio.strftime('%I:%M %p')
        bloques_en_hora = bloques.filter(hora_inicio=h_inicio)
        h_fin = bloques_en_hora.first().hora_fin.strftime('%I:%M %p') if bloques_en_hora.exists() else ''
        ws.cell(row=row_idx, column=1, value=f"{h_formatted} - {h_fin}").alignment = Alignment(horizontal='center')
        
        for p_dia in range(1, 6): # Lunes a Viernes
            bloque = bloques_en_hora.filter(dia_semana=str(p_dia)).first()
            if bloque:
                if bloque.tipo == 'CLASE':
                    texto = f"{bloque.asignatura.nombre if bloque.asignatura else ''}\n{bloque.docente.first_name if bloque.docente else ''}\nAula: {bloque.aula.nombre if bloque.aula else ''}"
                else:
                    texto = bloque.get_tipo_display()
                celda = ws.cell(row=row_idx, column=p_dia+1, value=texto)
                celda.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                # Optional coloring
                if bloque.tipo != 'CLASE':
                    celda.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        row_idx += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=horario_{horario.ano_grado}_{horario.seccion}.xlsx'
    wb.save(response)
    return response

def api_validar_horario(request):
    """Devuelve el ID del horario si existe para el periodo, año y sección dado"""
    periodo_id = request.GET.get('periodo_id')
    ano = request.GET.get('ano')
    seccion = request.GET.get('seccion')
    
    try:
        h = Horario.objects.get(periodo_id=periodo_id, ano_grado=ano, seccion=seccion)
        return JsonResponse({'status': 'ok', 'horario_id': h.id})
    except Horario.DoesNotExist:
         return JsonResponse({'status': 'not_found'})

@login_required
@csrf_exempt
def api_restablecer_horarios(request):
    """
    Elimina físicamente uno o varios horarios seleccionados.
    Solo permitido para el rol ADMINISTRATIVO.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    if request.user.rol not in ['ADMINISTRATIVO', 'PERSONAL', 'DESARROLLADOR']:
        return JsonResponse({'error': 'No tienes permisos para realizar esta acción.'}, status=403)
        
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'JSON inválido'}, status=400)
        
    horarios_ids = data.get('horarios_ids', [])
    if not horarios_ids or not isinstance(horarios_ids, list):
        return JsonResponse({'error': 'Debe proporcionar una lista de IDs de horarios válidos.'}, status=400)
        
    try:
        # Eliminar los horarios físicamente de la base de datos
        # Esto eliminará en cascada los BloqueHorario asociados
        eliminados, _ = Horario.objects.filter(id__in=horarios_ids).delete()
        return JsonResponse({
            'ok': True, 
            'mensaje': f'Se han eliminado {eliminados} horarios correctamente.'
        })
    except Exception as e:
        return JsonResponse({'error': f'Error al eliminar horarios: {str(e)}'}, status=500)


@login_required
@csrf_exempt
def api_limpiar_horario(request, horario_id):
    """
    Elimina todos los bloques de un horario para dejarlo en blanco.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
        
    try:
        horario = get_object_or_404(Horario, id=horario_id)
        # Auditoría
        from auditoria.models import registrar_evento
        registrar_evento(
            'INACTIVACION', 
            f"Horario dejado en blanco (se eliminaron todos sus bloques): ({horario.ano_grado}-{horario.seccion})", 
            'Horarios', 
            request.user.username, 
            'MEDIO'
        )
        # Eliminar todos los bloques del horario
        bloques_eliminados, _ = BloqueHorario.objects.filter(horario=horario).delete()
        return JsonResponse({
            'status': 'ok',
            'mensaje': f'Se han eliminado {bloques_eliminados} bloques. El horario está ahora en blanco.'
        })
    except Exception as e:
        logger.error(f"[Horarios] Error al dejar en blanco el horario: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def api_periodos(request):
    """Lista todos los períodos académicos ordenados por ID descendente."""
    periodos = list(
        PeriodoAcademico.objects.all().order_by('-activo', '-id').values('id', 'nombre', 'activo')
    )
    return JsonResponse({'periodos': periodos})


@login_required
@csrf_exempt
def api_crear_periodo(request):
    """Crea un nuevo período académico. Solo PERSONAL (superusuario/directora)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    if request.user.rol not in ['PERSONAL', 'DESARROLLADOR']:
        return JsonResponse({'error': 'No tienes permisos para crear períodos académicos.'}, status=403)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    nombre = data.get('nombre', '').strip()
    if not nombre:
        return JsonResponse({'error': 'El nombre del período no puede estar vacío.'}, status=400)

    if PeriodoAcademico.objects.filter(nombre=nombre).exists():
        return JsonResponse({'error': f'Ya existe un período con el nombre "{nombre}".'}, status=400)

    from django.utils import timezone
    periodo = PeriodoAcademico.objects.create(
        nombre=nombre,
        fecha_inicio=timezone.now().date(),
        fecha_fin=timezone.now().date(),
        activo=True,
    )

    # Auditoría
    try:
        from auditoria.models import registrar_evento
        registrar_evento(
            'CREACION',
            f'Período académico creado: "{nombre}"',
            'Horarios',
            request.user.username,
            'INFORMATIVO',
        )
    except Exception:
        pass

    return JsonResponse({
        'ok': True,
        'periodo': {'id': periodo.id, 'nombre': periodo.nombre, 'activo': periodo.activo}
    })


@login_required
def api_periodos(request):
    """Lista todos los períodos académicos ordenados por ID descendente."""
    periodos = list(
        PeriodoAcademico.objects.all().order_by('-activo', '-id').values('id', 'nombre', 'activo')
    )
    return JsonResponse({'periodos': periodos})


@login_required
@csrf_exempt
def api_crear_periodo(request):
    """Crea un nuevo período académico. Solo PERSONAL (superusuario/directora)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    if request.user.rol not in ['PERSONAL', 'DESARROLLADOR']:
        return JsonResponse({'error': 'No tienes permisos para crear períodos académicos.'}, status=403)

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    nombre = data.get('nombre', '').strip()
    if not nombre:
        return JsonResponse({'error': 'El nombre del período no puede estar vacío.'}, status=400)

    if PeriodoAcademico.objects.filter(nombre=nombre).exists():
        return JsonResponse({'error': f'Ya existe un período con el nombre "{nombre}".'}, status=400)

    from django.utils import timezone
    periodo = PeriodoAcademico.objects.create(
        nombre=nombre,
        fecha_inicio=timezone.now().date(),
        fecha_fin=timezone.now().date(),
        activo=True,
    )

    # Auditoría
    try:
        from auditoria.models import registrar_evento
        registrar_evento(
            'CREACION',
            f'Período académico creado: "{nombre}"',
            'Horarios',
            request.user.username,
            'INFORMATIVO',
        )
    except Exception:
        pass

    return JsonResponse({
        'ok': True,
        'periodo': {'id': periodo.id, 'nombre': periodo.nombre, 'activo': periodo.activo}
    })

def api_asignaciones_horario(request):
    """Devuelve las asignaciones cruzadas de docentes y materias para un año, sección y periodo específico."""
    periodo_id = request.GET.get('periodo_id')
    ano_grado = request.GET.get('ano_grado')
    seccion = request.GET.get('seccion')
    
    if not all([periodo_id, ano_grado, seccion]):
        return JsonResponse([], safe=False)
        
    try:
        from docentes.models import AsignacionDocente
        asignaciones = AsignacionDocente.objects.filter(
            periodo_id=periodo_id, 
            ano_grado=ano_grado, 
            seccion=seccion, 
            activa=True
        ).select_related('docente', 'asignatura')
        
        data = []
        for a in asignaciones:
            # Usar nombre completo del usuario si existe, si no username
            doc_nombre = a.docente.nombre_completo or f"{a.docente.first_name} {a.docente.last_name}".strip()
            if not doc_nombre:
                doc_nombre = a.docente.username
                
            data.append({
                'docente_id': a.docente.id,
                'docente_nombre': doc_nombre,
                'asignatura_id': a.asignatura.id,
                'asignatura_nombre': a.asignatura.nombre,
            })
        return JsonResponse(data, safe=False)
    except Exception as e:
        logger.error(f"[Horarios] Error al consultar asignaciones: {e}")
        return JsonResponse([], safe=False)
