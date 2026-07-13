import openpyxl

class _RowWrapper:
    def __init__(self, data):
        self.data = data
    def __iter__(self):
        return iter(self.data)
    def __len__(self):
        return len(self.data)
    def __getitem__(self, idx):
        return self.data[idx]
    @property
    def values(self):
        return self.data
    @property
    def iloc(self):
        return self

class _ILocWrapper:
    def __init__(self, data):
        self.data = data
    def __getitem__(self, idx):
        if isinstance(idx, tuple):
            row_idx, col_idx = idx
            row = self.data[row_idx]
            if col_idx < len(row):
                return row[col_idx]
            return ""
        return _RowWrapper(self.data[idx])

class MockDataFrame:
    def __init__(self, data):
        self.data = data
        self.iloc = _ILocWrapper(data)
        self.columns = data[0] if data else []
        
    def __len__(self):
        return len(self.data)
        
    def fillna(self, value):
        new_data = []
        for row in self.data:
            new_row = []
            for v in row:
                if v is None:
                    new_row.append(value)
                elif str(v).lower() == 'nan':
                    new_row.append(value)
                else:
                    new_row.append(str(v))
            new_data.append(new_row)
        return MockDataFrame(new_data)

def read_excel(file_object, header=None, dtype=str):
    wb = openpyxl.load_workbook(file_object, data_only=True)
    ws = wb.active
    data = []
    max_col = 0
    # First pass to find max columns
    for row in ws.iter_rows(values_only=True):
        if len(row) > max_col:
            max_col = len(row)
            
    # Second pass to normalize rows
    for row in ws.iter_rows(values_only=True):
        row_list = list(row)
        while len(row_list) < max_col:
            row_list.append(None)
        data.append(row_list)
        
    return MockDataFrame(data)

def read_excel_all_sheets(file_object):
    """Lee todas las hojas de un archivo Excel y retorna un dict {nombre_hoja: MockDataFrame}."""
    wb = openpyxl.load_workbook(file_object, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = []
        max_col = 0
        # First pass to find max columns
        for row in ws.iter_rows(values_only=True):
            if len(row) > max_col:
                max_col = len(row)

        # Second pass to normalize rows
        for row in ws.iter_rows(values_only=True):
            row_list = list(row)
            while len(row_list) < max_col:
                row_list.append(None)
            data.append(row_list)

        sheets[sheet_name] = MockDataFrame(data)
    return sheets

class pd:
    @staticmethod
    def read_excel(file_object, header=None, dtype=str):
        return read_excel(file_object, header=header, dtype=dtype)

    @staticmethod
    def read_excel_all_sheets(file_object):
        return read_excel_all_sheets(file_object)


def reset_db_sequences():
    """
    Restablece todas las secuencias de claves primarias en bases de datos PostgreSQL
    para evitar errores de duplicación de claves (IntegrityError) causados por inserciones
    con IDs explícitos (por ejemplo, mediante scripts de importación de datos REST o manuales).
    """
    import logging
    logger = logging.getLogger(__name__)
    
    from django.db import connection
    if connection.vendor == 'postgresql':
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    DO $$
                    DECLARE
                        r RECORD;
                    BEGIN
                        FOR r IN
                            SELECT 
                                tc.table_name, 
                                cc.column_name, 
                                pg_get_serial_sequence(tc.table_schema || '.' || tc.table_name, cc.column_name) as seq_name
                            FROM information_schema.table_constraints tc
                            JOIN information_schema.constraint_column_usage ccu 
                                ON tc.constraint_name = ccu.constraint_name 
                                AND tc.table_schema = ccu.table_schema
                            JOIN information_schema.columns cc 
                                ON ccu.table_name = cc.table_name 
                                AND ccu.column_name = cc.column_name
                                AND ccu.table_schema = cc.table_schema
                            WHERE tc.constraint_type = 'PRIMARY KEY' 
                              AND tc.table_schema = 'public'
                              AND pg_get_serial_sequence(tc.table_schema || '.' || tc.table_name, cc.column_name) IS NOT NULL
                        LOOP
                            EXECUTE 'SELECT setval(' || quote_literal(r.seq_name) || ', COALESCE((SELECT MAX(' || quote_ident(r.column_name) || ') FROM ' || quote_ident(r.table_name) || '), 1))';
                        END LOOP;
                    END $$;
                """)
            logger.info("Secuencias de base de datos PostgreSQL sincronizadas con éxito.")
        except Exception as e:
            logger.error(f"Error al sincronizar las secuencias de base de datos: {e}", exc_info=True)

