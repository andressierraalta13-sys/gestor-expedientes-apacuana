import os
import re
import unicodedata
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.cell.cell import Cell
from django.conf import settings
from django.core.files.base import ContentFile
from django.template.loader import render_to_string
from xhtml2pdf import pisa

from estudiantes.models import Estudiante
from inscripciones.models import PeriodoAcademico, Inscripcion, Asignatura
from calificaciones.models import Calificacion, NotaCertificada

ASIGNATURAS_PREDEFINIDAS = [
    "LENGUA Y LITERATURA", "IDIOMAS", "MATEMÁTICA", "MATEMATICA",
    "EDUCACIÓN FÍSICA", "EDUCACION FISICA",
    "BIOLOGÍA, AMBIENTE Y TECNOLOGÍA", "BIOLOGIA, AMBIENTE Y TECNOLOGIA",
    "FÍSICA", "FISICA", "QUÍMICA", "QUIMICA",
    "GEOGRAFÍA, HISTORIA, Y SOBERANÍA NACIONAL", "GEOGRAFIA, HISTORIA Y SOBERANIA NACIONAL",
    "INNOVACIÓN TECNOLOGÍA Y PRODUCTIVIDAD", "INNOVACION TECNOLOGIA Y PRODUCTIVIDAD",
    "ORIENTACIÓN VOCACIONAL", "ORIENTACION VACACIONAL"
]

# ── Plantilla Excel oficial vigente ──────────────────────────────────────────
PLANTILLA_XLSX = 'FORMATO NC EJEMPLO.xlsx'
HOJA_XLSX = 'Hoja1'

# ── Datos institucionales de Apacuana ────────────────────────────────────────
PLANTEL = {
    'nombre': 'UNIDAD EDUCATIVA NACIONAL APACUANA',
    'eponimo': 'U. E. N. APACUANA',
    'codigo_plantel': 'OD24061508',
    'direccion': 'URBANIZACION CIUDAD MIRANDA TERCERA ETAPA MANZANA 80',
    'telefono': '04241324154',
    'municipio': 'CRISTOBAL ROJAS',
    'entidad_federal': 'MIRANDA',
    'zona_educativa': 'MIRANDA',
    'localidad': 'CHARALLAVE',
    'plan_estudio': 'EDUCACIÓN MEDIA GENERAL',
    'codigo_plan': '31060',
    'director_nombre': 'DORCA DIAZ',
    'director_cedula': 'V-18.930.481',
}

MESES = {
    '1': 'ENERO', '01': 'ENERO', '2': 'FEBRERO', '02': 'FEBRERO',
    '3': 'MARZO', '03': 'MARZO', '4': 'ABRIL', '04': 'ABRIL',
    '5': 'MAYO', '05': 'MAYO', '6': 'JUNIO', '06': 'JUNIO',
    '7': 'JULIO', '07': 'JULIO', '8': 'AGOSTO', '08': 'AGOSTO',
    '9': 'SEPTIEMBRE', '09': 'SEPTIEMBRE', '10': 'OCTUBRE',
    '11': 'NOVIEMBRE', '12': 'DICIEMBRE',
}

# 8 Materias Oficiales por año
MATERIAS_SISTEMA = [
    "LENGUA Y LITERATURA",
    "IDIOMAS",
    "MATEMÁTICA",
    "EDUCACIÓN FÍSICA",
    "BIOLOGÍA, AMBIENTE Y TECNOLOGÍA",
    "FÍSICA",
    "QUÍMICA",
    "GEOGRAFÍA, HISTORIA, Y SOBERANÍA NACIONAL",
]

AREAS_POR_ANO = {ano: MATERIAS_SISTEMA for ano in range(1, 6)}

_AREA_SINONIMOS = {
    "LENGUA Y LITERATURA": ["LENGUA Y LITERATURA", "CASTELLANO", "LENGUA", "LENGUAJE"],
    "IDIOMAS": ["IDIOMAS", "INGLÉS", "INGLES", "INGLÉS Y OTRAS LENGUAS EXTRANJERAS", "INGLES Y OTRAS LENGUAS EXTRANJERAS"],
    "MATEMÁTICA": ["MATEMÁTICA", "MATEMATICA", "MATEMÁTICAS", "MATEMATICAS"],
    "EDUCACIÓN FÍSICA": ["EDUCACIÓN FÍSICA", "EDUCACION FISICA", "EDUACIÓN FÍSICA", "EDUACION FISICA"],
    "BIOLOGÍA, AMBIENTE Y TECNOLOGÍA": [
        "BIOLOGÍA, AMBIENTE Y TECNOLOGÍA", "BIOLOGIA, AMBIENTE Y TECNOLOGIA",
        "BIOLOGÍA", "BIOLOGIA", "CIENCIAS BIOLOGICAS",
        "A.C.T", "CIENCIAS NATURALES", "AMBIENTE Y TECNOLOGÍA",
    ],
    "FÍSICA": ["FÍSICA", "FISICA"],
    "QUÍMICA": ["QUÍMICA", "QUIMICA"],
    "GEOGRAFÍA, HISTORIA, Y SOBERANÍA NACIONAL": [
        "GEOGRAFÍA, HISTORIA, Y SOBERANÍA NACIONAL",
        "GEOGRAFIA, HISTORIA, Y SOBERANIA NACIONAL",
        "GEOGRAFÍA, HISTORIA Y SOBERANÍA NACIONAL",
        "GEOGRAFIA, HISTORIA Y SOBERANIA NACIONAL",
        "GEOGRAFÍA, HISTORIA Y CIUDADANÍA",
        "GEOGRAFIA, HISTORIA Y CIUDADANIA",
        "GEOGRAFÍA/HISTORIA CIUDADANÍA",
        "GEOGRAFÍA HISTORIA CIUDADANÍA",
        "G.H.C", "F.S.N",
    ],
    "INNOVACIÓN TECNOLÓGICA Y PRODUCTIVA": [
        "INNOVACIÓN TECNOLÓGICA Y PRODUCTIVA", "INNOVACION TECNOLOGICA Y PRODUCTIVA",
        "INNOVACIÓN TECNOLOGÍA Y PRODUCTIVIDAD", "INNOVACION TECNOLOGIA Y PRODUCTIVIDAD",
        "I.T.P",
    ],
    "ORIENTACIÓN VOCACIONAL": [
        "ORIENTACIÓN VOCACIONAL", "ORIENTACION VOCACIONAL",
        "ORIENTACIÓN VACACIONAL", "ORIENTACION VACACIONAL",
        "ORIENTACIÓN Y CONVIVENCIA", "ORIENTACION Y CONVIVENCIA",
    ],
}

def convertir_nota_a_letras(nota):
    try:
        n = int(round(float(nota)))
    except (ValueError, TypeError):
        return ""
    
    letras = {
        1: "UNO", 2: "DOS", 3: "TRES", 4: "CUATRO", 5: "CINCO",
        6: "SEIS", 7: "SIETE", 8: "OCHO", 9: "NUEVE", 10: "DIEZ",
        11: "ONCE", 12: "DOCE", 13: "TRECE", 14: "CATORCE", 15: "QUINCE",
        16: "DIECISEIS", 17: "DIECISIETE", 18: "DIECIOCHO", 19: "DIECINUEVE", 20: "VEINTE"
    }
    return letras.get(n, "")

def _canon_asignatura(nombre):
    s = unicodedata.normalize('NFKD', str(nombre or '')).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r'[^A-Za-z0-9 ]', ' ', s).upper()
    s = re.sub(r'\s+', ' ', s).strip()
    s = s.replace('EDUACION', 'EDUCACION')
    if s.startswith('INNOVACION'):
        return 'INNOVACION'
    if s.startswith('ORIENTACION'):
        return 'ORIENTACION'
    return s

def _nota_para_area(area_nombre, notas_dict, usados=None):
    claves = [area_nombre] + _AREA_SINONIMOS.get(area_nombre, [])
    for clave in claves:
        canon = _canon_asignatura(clave)
        if not canon:
            continue
        if usados is not None and canon in usados:
            continue
        if canon in notas_dict:
            if usados is not None:
                usados.add(canon)
            return notas_dict[canon]
    return None

def _notas_definitivas_por_ano(estudiante, ano_grad):
    cals = Calificacion.objects.filter(
        inscripcion__estudiante=estudiante,
        inscripcion__ano_grado=ano_grad,
        tipo='DEF',
    )
    return {_canon_asignatura(c.asignatura.nombre): c.nota for c in cals}

def _entero_o_none(nota):
    try:
        return int(round(float(nota))) if nota is not None else None
    except (ValueError, TypeError):
        return None

def generar_nota_certificada_automatica(estudiante_id, usuario_nombre):
    """
    Genera el Excel oficial de Notas Certificadas con las 8 materias oficiales,
    formato de asteriscos exacto, áreas complementarias de Innovación y Orientación,
    y firmas oficiales.
    """
    try:
        estudiante = Estudiante.objects.get(cedula_identidad=estudiante_id)
    except Estudiante.DoesNotExist:
        raise Exception(f"No se encontró al estudiante con cédula {estudiante_id}")

    plantilla_path = os.path.join(settings.BASE_DIR, 'FORMATOS EXCEL', PLANTILLA_XLSX)
    if not os.path.exists(plantilla_path):
        raise Exception(f"No se encontró la plantilla oficial '{PLANTILLA_XLSX}'.")

    wb_out = openpyxl.load_workbook(plantilla_path)
    ws_out = wb_out[HOJA_XLSX] if HOJA_XLSX in wb_out.sheetnames else wb_out.active

    def _set(coord, valor):
        ws_out[coord] = valor

    # ── 1. Encabezado ─────────────────────────────────────────
    ahora = datetime.now()
    codigo_generado = f"{PLANTEL.get('codigo_plan', '31060')}"
    fecha_larga = f"{ahora.day:02d} DE {MESES.get(str(ahora.month), '').upper()} DE {ahora.year}"
    
    _set('R2', f"Código {codigo_generado}")
    _set('P3', f"{PLANTEL['localidad'].upper()}, {fecha_larga}")

    # ── 2. II. Datos del Plantel ──────────────────────────────
    _set('C6', PLANTEL['codigo_plantel'])
    _set('M6', PLANTEL['eponimo'])
    _set('C7', PLANTEL['direccion'])
    _set('Q7', PLANTEL['telefono'])
    _set('C8', PLANTEL['municipio'])
    _set('I8', PLANTEL['entidad_federal'])
    _set('Q8', PLANTEL['zona_educativa'])

    # ── 3. III. Datos del Estudiante ──────────────────────────
    ci_fmt = f"{int(estudiante.cedula_identidad):,}".replace(',', '.') if str(estudiante.cedula_identidad).isdigit() else str(estudiante.cedula_identidad)
    _set('D10', ci_fmt)

    if estudiante.fecha_nacimiento:
        fecha_nac_txt = f"{estudiante.fecha_nacimiento.day:02d} DE {MESES.get(str(estudiante.fecha_nacimiento.month), '').upper()} DE {estudiante.fecha_nacimiento.year}"
    else:
        fecha_nac_txt = ''
    _set('N10', fecha_nac_txt)

    _set('B11', (estudiante.apellidos or '').upper())
    _set('N11', (estudiante.nombres or '').upper())
    _set('E12', (estudiante.pais_nacimiento or 'VENEZUELA').upper())
    _set('N12', (estudiante.estado_nacimiento or '').upper())
    _set('Q12', (estudiante.municipio_nacimiento or '').upper())

    # ── 4. IV. Instituciones Educativas ───────────────────────
    _set('B16', PLANTEL['nombre'])
    _set('F16', PLANTEL['localidad'])
    _set('I16', PLANTEL['entidad_federal'])

    _set('B17', '************************')
    _set('F17', '******************')
    _set('I17', '***')

    _set('N15', '************************')
    _set('Q15', '******************')
    _set('T15', '***')

    _set('N16', '************************')
    _set('Q16', '******************')
    _set('T16', '***')

    _set('N17', '************************')
    _set('Q17', '******************')
    _set('T17', '***')

    # ── 5. Fechas de culminación por año ───────────────────────
    fechas_culminacion = {
        1: (estudiante.mes_culminacion_1er_ano, estudiante.ano_culminacion_1er_ano),
        2: (estudiante.mes_culminacion_2do_ano, estudiante.ano_culminacion_2do_ano),
        3: (estudiante.mes_culminacion_3er_ano, estudiante.ano_culminacion_3er_ano),
        4: (estudiante.mes_culminacion_4to_ano, estudiante.ano_culminacion_4to_ano),
        5: (estudiante.mes_culminacion_5to_ano, estudiante.ano_culminacion_5to_ano),
    }

    # ── 6. V. Plan de Estudio (Calificaciones 1° a 5° año) ────
    COLS = {
        'L': {'area': 1,  'nota': 4,  'letra': 5,  'te': 7,  'mes': 8,  'ano': 9,  'plantel': 10},
        'R': {'area': 13, 'nota': 15, 'letra': 16, 'te': 18, 'mes': 19, 'ano': 20, 'plantel': 21},
    }

    bloques_anos = {
        1: {'lado': 'L', 'row_start': 22, 'row_end': 29},
        2: {'lado': 'R', 'row_start': 22, 'row_end': 29},
        3: {'lado': 'L', 'row_start': 33, 'row_end': 40},
        4: {'lado': 'R', 'row_start': 33, 'row_end': 40},
        5: {'lado': 'L', 'row_start': 44, 'row_end': 51},
    }

    for ano_grad, cfg in bloques_anos.items():
        col = COLS[cfg['lado']]
        notas_dict = _notas_definitivas_por_ano(estudiante, ano_grad)
        mes_val, ano_val = fechas_culminacion.get(ano_grad, ("", ""))
        usados = set()

        for idx, area_nombre in enumerate(MATERIAS_SISTEMA):
            row_idx = cfg['row_start'] + idx
            # Escribir el nombre del área
            ws_out.cell(row=row_idx, column=col['area']).value = area_nombre

            nota = _nota_para_area(area_nombre, notas_dict, usados)
            num = _entero_o_none(nota)

            if num is not None:
                ws_out.cell(row=row_idx, column=col['nota']).value = num
                ws_out.cell(row=row_idx, column=col['letra']).value = convertir_nota_a_letras(num)
                ws_out.cell(row=row_idx, column=col['te']).value = 'F'
                ws_out.cell(row=row_idx, column=col['mes']).value = str(mes_val).zfill(2) if mes_val else '07'
                ws_out.cell(row=row_idx, column=col['ano']).value = str(ano_val) if ano_val else str(ahora.year)
                ws_out.cell(row=row_idx, column=col['plantel']).value = '01'
            else:
                ws_out.cell(row=row_idx, column=col['nota']).value = '*****'
                ws_out.cell(row=row_idx, column=col['letra']).value = '****************'
                ws_out.cell(row=row_idx, column=col['te']).value = '****'
                ws_out.cell(row=row_idx, column=col['mes']).value = '****'
                ws_out.cell(row=row_idx, column=col['ano']).value = '****'
                ws_out.cell(row=row_idx, column=col['plantel']).value = '****'

    # ── 7. Innovación Tecnología y Productividad (Filas 44 a 48) ─
    for ano_grad in range(1, 6):
        row_idx = 44 + (ano_grad - 1)
        notas_dict = _notas_definitivas_por_ano(estudiante, ano_grad)
        mes_val, ano_val = fechas_culminacion.get(ano_grad, ("", ""))
        nota_itp = _nota_para_area("INNOVACIÓN TECNOLÓGICA Y PRODUCTIVA", notas_dict)
        num_itp = _entero_o_none(nota_itp)

        ws_out.cell(row=row_idx, column=15).value = ano_grad
        if num_itp is not None:
            ws_out.cell(row=row_idx, column=16).value = num_itp
            ws_out.cell(row=row_idx, column=17).value = convertir_nota_a_letras(num_itp)
            ws_out.cell(row=row_idx, column=18).value = 'F'
            ws_out.cell(row=row_idx, column=19).value = str(mes_val).zfill(2) if mes_val else '07'
            ws_out.cell(row=row_idx, column=20).value = str(ano_val) if ano_val else str(ahora.year)
            ws_out.cell(row=row_idx, column=21).value = 1
        else:
            ws_out.cell(row=row_idx, column=16).value = '********'
            ws_out.cell(row=row_idx, column=17).value = '******'
            ws_out.cell(row=row_idx, column=18).value = '****'
            ws_out.cell(row=row_idx, column=19).value = '****'
            ws_out.cell(row=row_idx, column=20).value = '******'
            ws_out.cell(row=row_idx, column=21).value = '**'

    # ── 8. Orientación Vacacional (Filas 50 a 54) ─────────────
    for ano_grad in range(1, 6):
        row_idx = 50 + (ano_grad - 1)
        notas_dict = _notas_definitivas_por_ano(estudiante, ano_grad)
        nota_ov = _nota_para_area("ORIENTACIÓN VOCACIONAL", notas_dict)

        ws_out.cell(row=row_idx, column=15).value = ano_grad
        if nota_ov is not None:
            ws_out.cell(row=row_idx, column=16).value = str(nota_ov).strip().upper() or 'D'
        else:
            ws_out.cell(row=row_idx, column=16).value = '***************************'

    # ── 9. VI. Observaciones (A52) ────────────────────────────
    ws_out['A52'] = 'VI. Observaciones:'

    # ── 10. Firmas y Sellos (VII y VIII) ──────────────────────
    ws_out['A58'] = PLANTEL['director_nombre']
    ws_out['A60'] = PLANTEL['director_cedula']
    ws_out['E56'] = 'SELLO'

    ws_out['M58'] = '*************************'
    ws_out['M60'] = '***************'
    ws_out['P56'] = 'SELLO'

    # ── 11. Serializar a memoria ──────────────────────────────
    out_stream = BytesIO()
    wb_out.save(out_stream)
    out_stream.seek(0)
    xlsx_bytes = out_stream.read()

    nombre_archivo = f"NC_Auto_{estudiante.cedula_identidad}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    ya_existe = NotaCertificada.objects.filter(cedula_normalizada=estudiante.cedula_identidad).first()
    nota_obj = ya_existe or NotaCertificada()

    nota_obj.cedula_normalizada = estudiante.cedula_identidad
    nota_obj.nombre_completo = f"{estudiante.apellidos or ''} {estudiante.nombres or ''}".strip()
    nota_obj.nombres = estudiante.nombres or ''
    nota_obj.apellidos = estudiante.apellidos or ''
    nota_obj.cargado_por = f"{usuario_nombre} (Auto XLSX)"
    nota_obj.nombre_archivo_original = nombre_archivo
    nota_obj.save()

    return nota_obj, xlsx_bytes

def link_callback(uri, rel):
    sUrl = settings.STATIC_URL
    sRoot = settings.STATIC_ROOT
    mUrl = settings.MEDIA_URL
    mRoot = settings.MEDIA_ROOT

    if uri.startswith(mUrl):
        path = os.path.join(mRoot, uri.replace(mUrl, ""))
    elif uri.startswith(sUrl):
        path = os.path.join(sRoot, uri.replace(sUrl, ""))
    else:
        return uri

    if not os.path.isfile(path):
        raise Exception('URI no encontrada en el sistema de archivos: %s' % path)
    return path

def generar_nota_certificada_pdf_automatica(estudiante_id, usuario_nombre):
    """
    Genera el archivo PDF automáticamente a partir del Excel oficial completado.
    """
    nota_obj, xlsx_bytes = generar_nota_certificada_automatica(estudiante_id, usuario_nombre)
    
    # Exportar Excel a PDF mediante COM en Windows o fallback HTML
    try:
        import tempfile
        import subprocess
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f_xlsx:
            f_xlsx.write(xlsx_bytes)
            tmp_xlsx = f_xlsx.name
        
        tmp_pdf = tmp_xlsx.replace('.xlsx', '.pdf')
        
        ps_cmd = f'''
        $excel = New-Object -ComObject Excel.Application
        $excel.Visible = $false
        $excel.DisplayAlerts = $false
        $wb = $excel.Workbooks.Open("{tmp_xlsx}")
        $ws = $wb.Sheets.Item(1)
        $ws.PageSetup.PaperSize = 1
        $ws.PageSetup.PrintArea = '$A$1:$U$63'
        $ws.PageSetup.TopMargin = 2.85
        $ws.PageSetup.BottomMargin = 2.85
        $ws.PageSetup.LeftMargin = 2.85
        $ws.PageSetup.RightMargin = 2.85
        $ws.PageSetup.CenterHorizontally = $true
        $ws.PageSetup.CenterVertically = $false
        $ws.PageSetup.Zoom = $false
        $ws.PageSetup.FitToPagesWide = 1
        $ws.PageSetup.FitToPagesTall = 1
        $wb.ExportAsFixedFormat(0, "{tmp_pdf}")
        $wb.Close($false)
        $excel.Quit()
        [System.Runtime.Interopservices.Marshal]::ReleaseComObject($excel) | Out-Null
        '''
        
        subprocess.run(["powershell", "-ExecutionPolicy", "Bypass", "-Command", ps_cmd], check=True)
        
        with open(tmp_pdf, 'rb') as f_pdf:
            pdf_bytes = f_pdf.read()
            
        try:
            os.remove(tmp_xlsx)
            os.remove(tmp_pdf)
        except Exception:
            pass
            
        return nota_obj, pdf_bytes
    except Exception as e:
        # Fallback a HTML rendering si no está disponible COM
        pass

    try:
        estudiante = Estudiante.objects.get(cedula_identidad=estudiante_id)
    except Estudiante.DoesNotExist:
        raise Exception(f"No se encontró al estudiante con cédula {estudiante_id}")
        
    ahora = datetime.now()
    codigo_generado = f"{PLANTEL.get('codigo_plan', '31060')}"
    fecha_larga = f"{ahora.day:02d} DE {MESES.get(str(ahora.month), '').upper()} DE {ahora.year}"
    fecha_nac_str = estudiante.fecha_nacimiento.strftime('%d/%m/%Y') if estudiante.fecha_nacimiento else ''

    fechas_culminacion = {
        1: (estudiante.mes_culminacion_1er_ano, estudiante.ano_culminacion_1er_ano),
        2: (estudiante.mes_culminacion_2do_ano, estudiante.ano_culminacion_2do_ano),
        3: (estudiante.mes_culminacion_3er_ano, estudiante.ano_culminacion_3er_ano),
        4: (estudiante.mes_culminacion_4to_ano, estudiante.ano_culminacion_4to_ano),
        5: (estudiante.mes_culminacion_5to_ano, estudiante.ano_culminacion_5to_ano)
    }

    anos_nombres = {1: "PRIMER", 2: "SEGUNDO", 3: "TERCER", 4: "CUARTO", 5: "QUINTO"}
    anios = []

    for ano_grad in range(1, 6):
        notas_dict = _notas_definitivas_por_ano(estudiante, ano_grad)
        mes_val, ano_val = fechas_culminacion.get(ano_grad, ("", ""))
        usados = set()

        filas = []
        for area in AREAS_POR_ANO[ano_grad]:
            num = _entero_o_none(_nota_para_area(area, notas_dict, usados))
            filas.append({
                'area': area,
                'num': num if num is not None else '',
                'letras': convertir_nota_a_letras(num) if num is not None else '',
            })

        anios.append({
            'nombre': anos_nombres[ano_grad],
            'mes': str(mes_val).zfill(2) if mes_val else '',
            'ano': ano_val or '',
            'filas': filas,
        })

    def _logo(nombre):
        ruta = os.path.join(settings.MEDIA_ROOT, nombre)
        if not os.path.exists(ruta):
            return None
        return ruta.replace('\\', '/') if os.name == 'nt' else ruta

    context = {
        'plantel': PLANTEL,
        'codigo': codigo_generado,
        'fecha_larga': fecha_larga,
        'telefono': estudiante.telefono_representante or PLANTEL['telefono'],
        'e_cedula': estudiante.cedula_identidad,
        'fecha_nacimiento': fecha_nac_str,
        'apellido': estudiante.apellidos or '',
        'nombre': estudiante.nombres or '',
        'lugar_nacimiento': estudiante.lugar_nacimiento or '',
        'estado_nacimiento': estudiante.estado_nacimiento or '',
        'municipio_nacimiento': estudiante.municipio_nacimiento or '',
        'pais_nacimiento': estudiante.pais_nacimiento or 'Venezuela',
        'anios': anios,
        'logo_gobierno': _logo('gobierno.png'),
    }

    html_string = render_to_string('calificaciones/nota_certificada_pdf.html', context)
    
    out_stream = BytesIO()
    pisa_status = pisa.CreatePDF(
        html_string, 
        dest=out_stream, 
        link_callback=link_callback
    )
    
    if pisa_status.err:
        raise Exception('Hubo un error al generar el PDF de Notas Certificadas.')
        
    out_stream.seek(0)
    pdf_bytes = out_stream.read()
    
    return nota_obj, pdf_bytes

def procesar_nota_certificada(archivo_subido, usuario_nombre):
    archivo_subido.seek(0)
    try:
        wb_in = openpyxl.load_workbook(archivo_subido, data_only=True)
    except Exception as e:
        raise Exception(f"No se pudo leer el archivo Excel. Asegúrate de que sea un formato válido. ({e})")
        
    sheet_name = None
    for name in wb_in.sheetnames:
        if name.upper() in ["NCF", "HOJA1", "SHEET1"]:
            sheet_name = name
            break
            
    if not sheet_name:
        sheet_name = wb_in.sheetnames[0]
        
    ws_in = wb_in[sheet_name]

    # Extraer Cédula de D10 o E11
    cedula_raw = str(ws_in['D10'].value or ws_in['E11'].value or '')
    cedula = re.sub(r'\D', '', cedula_raw)
    if not cedula or len(cedula) < 6:
        raise Exception(f"No se encontró una cédula válida en el archivo. Valor encontrado: '{cedula_raw}'")

    try:
        estudiante = Estudiante.objects.get(cedula_identidad=cedula)
    except Estudiante.DoesNotExist:
        raise Exception(f"No se encontró al estudiante con cédula {cedula} en la base de datos.")

    periodo = PeriodoAcademico.objects.filter(activo=True).first()
    if not periodo:
        periodo = PeriodoAcademico.objects.first()

    # Guardar en NotaCertificada
    nombre_archivo = getattr(archivo_subido, 'name', f"NC_{cedula}.xlsx")
    ya_existe = NotaCertificada.objects.filter(cedula_normalizada=cedula).first()
    nota_obj = ya_existe or NotaCertificada()

    nota_obj.cedula_normalizada = cedula
    nota_obj.nombre_completo = f"{estudiante.apellidos or ''} {estudiante.nombres or ''}".strip()
    nota_obj.nombres = estudiante.nombres or ''
    nota_obj.apellidos = estudiante.apellidos or ''
    nota_obj.cargado_por = usuario_nombre
    nota_obj.nombre_archivo_original = nombre_archivo
    nota_obj.save()

    return nota_obj
