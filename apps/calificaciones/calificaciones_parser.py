"""
Parser especializado para el formato de calificaciones multi-hoja.
Lee un archivo Excel con 3 hojas (M1, M2, M3) que corresponden a los
lapsos académicos L1, L2 y L3 respectivamente.

Estructura esperada por hoja:
  Filas 1-6:  Encabezado institucional y metadatos
  Fila 7:     Encabezado de columnas (10 materias en cols D-M)
  Fila 11+:   Datos de estudiantes (cédula en col B, notas en D-M)

Autor: Gestor Apacuana
"""

import re
import logging
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)

# ── Mapeo: nombre en el encabezado Excel → nombre canónico en Asignatura.nombre ──
# Las claves se normalizan (mayúsculas, sin tildes, sin saltos de línea)
# para hacer match flexible.
MATERIA_MAP = {
    'CASTELLANO':           'LENGUA Y LITERATURA',
    'LENGUA':               'LENGUA Y LITERATURA',
    'LENGUA Y LITERATURA':  'LENGUA Y LITERATURA',
    'IDIOMAS':              'IDIOMAS',
    'INGLES':               'IDIOMAS',
    'MATEMATICAS':          'MATEMÁTICA',
    'MATEMATICA':           'MATEMÁTICA',
    'EDUCACION FISICA':     'EDUCACIÓN FÍSICA',
    'A.C.T.':               'BIOLOGIA, AMBIENTE Y TECNOLOGIA',
    'ACT':                  'BIOLOGIA, AMBIENTE Y TECNOLOGIA',
    'BIOLOGIA':             'BIOLOGIA, AMBIENTE Y TECNOLOGIA',
    'FISICA':               'FÍSICA',
    'QUIMICA':              'QUÍMICA',
    'GEOGRAFIA':            'GEOGRAFÍA, HISTORIA , Y SOBERANÍA NACIONAL',
    'HISTORIA':             'GEOGRAFÍA, HISTORIA , Y SOBERANÍA NACIONAL',
    'CIUDADANIA':           'GEOGRAFÍA, HISTORIA , Y SOBERANÍA NACIONAL',
    'GHC':                  'GEOGRAFÍA, HISTORIA , Y SOBERANÍA NACIONAL',
    'G.H.C':                'GEOGRAFÍA, HISTORIA , Y SOBERANÍA NACIONAL',
    'ORIENTACION VOCACIONAL': 'ORIENTACIÓN VOCACIONAL',
    'ORIENTACION':          'ORIENTACIÓN VOCACIONAL',
    'I.T.P.':               'INNOVACIÓN TECNOLÓGICA Y PRODUCTIVA',
    'ITP':                  'INNOVACIÓN TECNOLÓGICA Y PRODUCTIVA',
    'INNOVACION':           'INNOVACIÓN TECNOLÓGICA Y PRODUCTIVA',
}

# Mapeo de hojas a tipos de lapso
HOJA_LAPSO_MAP = {
    'M1': 'L1',
    'M2': 'L2',
    'M3': 'L3',
}

# Columnas fijas del formato (1-indexed)
COL_NRO     = 1   # A
COL_CEDULA  = 2   # B
COL_NOMBRE  = 3   # C
COLS_NOTAS  = list(range(4, 14))  # D=4 hasta M=13 (10 materias)
FILA_HEADER = 7   # Fila del encabezado de columnas
FILA_DATOS  = 11  # Primera fila de datos de estudiantes

# Metadatos del formato
FILA_META = 4
COL_ANO = 2       # B4: año calendario
COL_SECCION = 4   # D4: sección
COL_ANO_ESCOLAR = 7  # G4: año escolar (1DO, 2DO, etc.)

# Patrones de año escolar
YEAR_PATTERNS = {
    '1ER': 1, '1RO': 1, 'PRIMERO': 1, '1DO': 1,
    '2DO': 2, '2NDO': 2, 'SEGUNDO': 2,
    '3ER': 3, '3RO': 3, 'TERCERO': 3,
    '4TO': 4, 'CUARTO': 4,
    '5TO': 5, 'QUINTO': 5,
}


def normalizar_cedula(valor) -> str:
    """
    Normaliza una cédula eliminando todos los caracteres no numéricos.
    '13.094.302' → '13094302'
    '13094302'   → '13094302'
    'V-13.094.302' → '13094302'
    """
    if valor is None:
        return ''
    return re.sub(r'\D', '', str(valor).strip())


def _normalizar_header(texto: str) -> str:
    """Normaliza texto de encabezado para matching: mayúsculas, sin tildes, sin \\n."""
    if not texto:
        return ''
    texto = str(texto).upper().strip()
    texto = texto.replace('\n', ' ').replace('\r', '')
    # Eliminar tildes
    reemplazos = {'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ñ': 'N'}
    for k, v in reemplazos.items():
        texto = texto.replace(k, v)
    # Comprimir espacios múltiples
    texto = re.sub(r'\s+', ' ', texto).strip()
    return texto


def _safe_float(val) -> Optional[float]:
    """Convierte un valor a float válido en rango escolar (0-20)."""
    if val is None:
        return None
    try:
        f = float(str(val).replace(',', '.').strip())
        if f != f:  # NaN
            return None
        if not (0.0 <= f <= 20.0):
            return None
        return round(f, 2)
    except (ValueError, TypeError):
        return None


def _detectar_ano_escolar(ws) -> Optional[int]:
    """Extrae el año escolar (1-5) de la fila de metadatos."""
    val = ws.cell(row=FILA_META, column=COL_ANO_ESCOLAR).value
    if val is None:
        return None
    texto = str(val).upper().strip()
    for pattern, grado in YEAR_PATTERNS.items():
        if pattern in texto:
            return grado
    # Intentar como número directo
    try:
        n = int(float(texto))
        if 1 <= n <= 5:
            return n
    except (ValueError, TypeError):
        pass
    return None


def _detectar_seccion(ws) -> str:
    """Extrae la sección de la fila de metadatos."""
    val = ws.cell(row=FILA_META, column=COL_SECCION).value
    if val is None:
        return ''
    return str(val).strip().upper()[:1]


def _construir_mapa_materias(ws) -> dict:
    """
    Lee la fila de encabezado y construye un mapeo:
    {col_index: nombre_canonico_asignatura}
    """
    mapa = {}
    for col_idx in COLS_NOTAS:
        raw = ws.cell(row=FILA_HEADER, column=col_idx).value
        if raw is None:
            continue
        header_norm = _normalizar_header(raw)

        # Intentar match directo
        if header_norm in MATERIA_MAP:
            mapa[col_idx] = MATERIA_MAP[header_norm]
            continue

        # Intentar match parcial (la primera palabra clave que coincida)
        encontrado = False
        for alias, canonico in MATERIA_MAP.items():
            alias_norm = _normalizar_header(alias)
            if alias_norm in header_norm or header_norm in alias_norm:
                mapa[col_idx] = canonico
                encontrado = True
                break

        if not encontrado:
            logger.warning(
                f"[CalificacionesParser] Columna {col_idx} con encabezado "
                f"'{raw}' no mapeada a ninguna materia conocida."
            )

    return mapa


@dataclass
class EstudianteCalificaciones:
    """Almacena las calificaciones de un estudiante agrupadas por lapso."""
    cedula_raw: str
    cedula_normalizada: str
    nombre: str
    notas: dict = field(default_factory=dict)
    # notas = { 'L1': {'LENGUA Y LITERATURA': 13, 'IDIOMAS': 14, ...},
    #           'L2': {...}, 'L3': {...} }
    errores: list = field(default_factory=list)


@dataclass
class CalificacionesParserResult:
    """Resultado del parsing del archivo de calificaciones."""
    estudiantes: dict = field(default_factory=dict)
    # {cedula_normalizada: EstudianteCalificaciones}
    ano_escolar: Optional[int] = None
    seccion: str = ''
    hojas_procesadas: list = field(default_factory=list)
    errores_globales: list = field(default_factory=list)
    advertencias: list = field(default_factory=list)
    total_notas_leidas: int = 0
    materias_detectadas: list = field(default_factory=list)


class CalificacionesParser:
    """
    Parser para el formato de calificaciones de Apacuana.
    Lee las hojas M1, M2, M3 (= Lapso 1, 2, 3) y extrae las notas
    de múltiples estudiantes por cédula.
    """

    def __init__(self, file_object):
        self.file_object = file_object
        self.result = CalificacionesParserResult()

    def parse(self) -> CalificacionesParserResult:
        """Punto de entrada principal."""
        try:
            wb = openpyxl.load_workbook(self.file_object, data_only=True)
        except Exception as e:
            self.result.errores_globales.append(f"Error al abrir el archivo Excel: {e}")
            return self.result

        # Verificar que existan las hojas esperadas
        hojas_disponibles = wb.sheetnames
        hojas_a_procesar = []

        for hoja_nombre, lapso in HOJA_LAPSO_MAP.items():
            if hoja_nombre in hojas_disponibles:
                hojas_a_procesar.append((hoja_nombre, lapso))
            else:
                self.result.advertencias.append(
                    f"Hoja '{hoja_nombre}' no encontrada en el archivo. "
                    f"Las calificaciones del {lapso} no serán procesadas."
                )

        if not hojas_a_procesar:
            self.result.errores_globales.append(
                "El archivo no contiene ninguna de las hojas esperadas (M1, M2, M3). "
                "Asegúrese de usar el formato de calificaciones correcto."
            )
            return self.result

        # Procesar cada hoja
        for hoja_nombre, lapso in hojas_a_procesar:
            ws = wb[hoja_nombre]
            self._procesar_hoja(ws, hoja_nombre, lapso)
            self.result.hojas_procesadas.append(hoja_nombre)

            # Extraer metadatos de la primera hoja procesada
            if self.result.ano_escolar is None:
                self.result.ano_escolar = _detectar_ano_escolar(ws)
            if not self.result.seccion:
                self.result.seccion = _detectar_seccion(ws)

        wb.close()
        return self.result

    def _procesar_hoja(self, ws, hoja_nombre: str, lapso: str):
        """Procesa una hoja individual del Excel."""
        mapa_materias = _construir_mapa_materias(ws)

        if not mapa_materias:
            self.result.advertencias.append(
                f"Hoja '{hoja_nombre}': no se detectaron columnas de materias "
                f"en la fila de encabezado ({FILA_HEADER})."
            )
            return

        # Registrar materias detectadas (solo la primera vez)
        if not self.result.materias_detectadas:
            self.result.materias_detectadas = sorted(set(mapa_materias.values()))

        # Recorrer filas de datos
        for row_idx in range(FILA_DATOS, ws.max_row + 1):
            cedula_raw = ws.cell(row=row_idx, column=COL_CEDULA).value
            if cedula_raw is None:
                continue

            cedula_norm = normalizar_cedula(cedula_raw)
            if not cedula_norm or len(cedula_norm) < 6:
                continue  # Fila vacía o cédula inválida

            nombre_raw = ws.cell(row=row_idx, column=COL_NOMBRE).value
            nombre = str(nombre_raw).strip() if nombre_raw else ''

            # Obtener o crear el registro del estudiante
            if cedula_norm not in self.result.estudiantes:
                self.result.estudiantes[cedula_norm] = EstudianteCalificaciones(
                    cedula_raw=str(cedula_raw).strip(),
                    cedula_normalizada=cedula_norm,
                    nombre=nombre,
                )

            est = self.result.estudiantes[cedula_norm]

            # Inicializar lapso si no existe
            if lapso not in est.notas:
                est.notas[lapso] = {}

            # Leer notas de cada materia
            for col_idx, materia_canonica in mapa_materias.items():
                val = ws.cell(row=row_idx, column=col_idx).value
                nota = _safe_float(val)
                if nota is not None:
                    est.notas[lapso][materia_canonica] = nota
                    self.result.total_notas_leidas += 1
