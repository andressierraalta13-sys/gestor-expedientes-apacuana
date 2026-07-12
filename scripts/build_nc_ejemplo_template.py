"""
Genera la plantilla LIMPIA de Notas Certificadas 'FORMATO NC EJEMPLO.xlsx' a
partir del formato oficial de referencia 'EJEMPLO.xlsx' (hoja 'NCF').

EJEMPLO.xlsx es un certificado real (relleno con datos de un estudiante y con
fórmulas de referencia externa `=[1]...`). Este script produce una versión
REUTILIZABLE que conserva EXACTAMENTE su diseño (bordes, celdas combinadas,
secciones "Epónimo", "Centro de Desarrollo", "COMPONENTES DE FORMACIÓN...",
firmas VII/VIII, etc.) pero:

  1. Elimina las imágenes sueltas y el enlace externo (el logo se reinyecta al
     generar cada certificado).
  2. Corrige el ÚNICO defecto estructural del original: en el bloque 1º/2º la
     cabecera de las columnas T-E y PLANTEL abarcaba 3 filas (20-22) mientras
     que el resto abarcaba 2 (20-21), lo que dejaba a la 1ª materia (LENGUA) sin
     casillas de T-E/Fecha/Plantel. Se recorta a 2 filas y se mueven las
     etiquetas Mes/Año de la fila 22 a la 21, de modo que las 10 materias del
     año tengan las 10 filas de datos completas (22-31).
  3. Completa el bloque de 5º año (izq.): las filas 56 y 57 estaban en blanco;
     se les asignan las materias INNOVACIÓN y ORIENTACIÓN con el mismo estilo y
     combinaciones que el resto, para que 5º también tenga sus 10 materias.
  4. Vacía todas las celdas de DATOS (calificaciones, datos del estudiante y del
     plantel, componentes) dejando solo las etiquetas: el generador las rellena.

Uso:  python scripts/build_nc_ejemplo_template.py
"""
import copy
import os

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(BASE_DIR, "FORMATOS EXCEL", "EJEMPLO.xlsx")
DST = os.path.join(BASE_DIR, "FORMATOS EXCEL", "FORMATO NC EJEMPLO.xlsx")
HOJA = "NCF"

# Materias en el orden EXACTO del formato (usadas solo para completar las filas
# 56/57 de 5º año que el original dejó en blanco).
SUBJECTS = [
    "LENGUA Y LITERATURA",
    "IDIOMAS",
    "MATEMÁTICA",
    "EDUCACIÓN FÍSICA",
    "BIOLOGÍA, AMBIENTE Y TECNOLOGÍA",
    "FÍSICA",
    "QUÍMICA",
    "GEOGRAFÍA, HISTORIA Y SOBERANÍA NACIONAL",
    "INNOVACIÓN TECNOLÓGICA Y PRODUCTIVA",
    "ORIENTACIÓN VOCACIONAL",
]

wb = openpyxl.load_workbook(SRC)
ws = wb[HOJA]

# ── 1. Imágenes y enlaces externos ────────────────────────────────────────────
ws._images = []
wb._external_links = []


def unmerge_if(ref):
    for mc in list(ws.merged_cells.ranges):
        if str(mc) == ref:
            ws.unmerge_cells(ref)
            return True
    return False


def blank(cell):
    """Vacía una celda; ignora las combinadas no-ancla (son de solo lectura)."""
    if not isinstance(cell, MergedCell):
        cell.value = None


def copy_style(dst_coord, src_coord, value="__keep__"):
    s = ws[src_coord]
    d = ws[dst_coord]
    d._style = copy.copy(s._style)
    d.number_format = s.number_format
    if value != "__keep__":
        d.value = value


# ── 2. Corregir cabecera del bloque 1º/2º (3 filas → 2 filas) ─────────────────
# Recorta las combinaciones de T-E y PLANTEL de 20-22 a 20-21 y sube Mes/Año.
for ref, new in (("H20:H22", "H20:H21"), ("K20:L22", "K20:L21"),
                 ("S20:S22", "S20:S21"), ("V20:V22", "V20:V21")):
    if unmerge_if(ref):
        ws.merge_cells(new)

# Etiquetas Mes/Año: de fila 22 (donde estorbaban a la 1ª materia) a fila 21.
for lbl_col, txt in (("I", "Mes"), ("J", "Año"), ("T", "Mes"), ("U", "Año")):
    copy_style(f"{lbl_col}21", f"{lbl_col}22")           # estilo de cabecera
    ws[f"{lbl_col}21"] = txt

# Al recortar la cabecera, la 1ª fila de datos (22 = LENGUA) perdió los bordes de
# las columnas T-E/Fecha/Plantel. Se clona el estilo de la 2ª fila (23 = IDIOMAS)
# en las columnas de datos para que LENGUA quede idéntica al resto de materias.
for c in list(range(5, 13)) + list(range(16, 23)):  # E..L (izq) y P..V (der)
    copy_style(f"{get_column_letter(c)}22", f"{get_column_letter(c)}23")

# ── 3. Completar 5º año (izq.): filas 56 y 57 (INNOVACIÓN, ORIENTACIÓN) ────────
# Se clonan estilo y combinaciones de la fila 55 (última materia con formato).
for r in (56, 57):
    for c in range(2, 12):  # B..K
        copy_style(f"{get_column_letter(c)}{r}", f"{get_column_letter(c)}55")
    for a, b in (("B", "D"), ("F", "G"), ("K", "L")):
        rng = f"{a}{r}:{b}{r}"
        if rng not in [str(m) for m in ws.merged_cells.ranges]:
            ws.merge_cells(rng)
ws["B56"] = SUBJECTS[8]   # INNOVACIÓN TECNOLÓGICA Y PRODUCTIVA
ws["B57"] = SUBJECTS[9]   # ORIENTACIÓN VOCACIONAL

# ── 3b. Normalizar los nombres de materia en TODOS los bloques ─────────────────
# El original mezcla nombres sin acentos y con erratas ("EDUACION FISICA",
# "INNOVACION TECNOLOGIA Y PRODUCTIVIDAD"). Se reescriben las 10 áreas de cada
# año, en su orden fijo, con la ortografía correcta y uniforme.
AREA_COL = {"L": 2, "R": 14}  # B (izq) / N (der)
for lado, r0, _r1 in (("L", 22, 31), ("R", 22, 31), ("L", 35, 44),
                      ("R", 35, 44), ("L", 48, 57)):
    for i in range(10):
        ws.cell(row=r0 + i, column=AREA_COL[lado]).value = SUBJECTS[i]

# ── 3c. Correcciones ortográficas de etiquetas estructurales ─────────────────
ETIQUETAS = {
    "L2": "CERTIFICACIÓN DE CALIFICACIONES EMG",
    "L3": "I. Plan de Estudio: EDUCACIÓN MEDIA GENERAL",
    "B6": ("II. Datos del Plantel o Centro de Desarrollo para la Calidad "
           "Educativa que Emite la Certificación:"),
    "B10": "III. Datos de Identificación del Estudiante:",
    "N45": "COMPONENTES DE FORMACIÓN CIENTÍFICA, TECNOLÓGICA Y PRODUCTIVA",
    "N46": "ÁREAS DE FORMACIÓN",
    "N48": "INNOVACIÓN TECNOLÓGICA Y PRODUCTIVA",
    "N53": "ÁREAS DE FORMACIÓN",
    "N55": "PARTICIPACIÓN EN GRUPOS DE CREACIÓN, RECREACIÓN Y PRODUCCIÓN",
    "B62": "VII. PLANTEL",
}
for coord, txt in ETIQUETAS.items():
    ws[coord] = txt

# ── 4. Vaciar TODAS las celdas de datos (deja solo etiquetas) ─────────────────
# 4a. Cualquier fórmula (externa `=[1]...` o interna de LETRAS) -> vacío.
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            cell.value = None

# 4b. Datos del plantel / estudiante / fecha (valores del certificado original).
for coord in ("S3", "Q4", "D7", "K7", "D8", "R8", "D9", "J9", "R9",
              "E11", "P11", "C12", "O12", "F13", "K13", "R13",
              "B65", "B67"):
    blank(ws[coord])

# 4c. Planteles (IV): nombres/localidades/EF del certificado original.
for coord in ("C16", "G16", "K16", "C17", "G17", "K17",
              "O15", "R15", "V15", "O16", "R16", "V16", "O17", "R17", "V17"):
    blank(ws[coord])

# 4d. Calificaciones de cada bloque de año (nota, letras, T-E, mes, año, plantel).
COLS = {
    "L": {"nota": 5, "letra": 6, "te": 8, "mes": 9, "ano": 10, "plantel": 11},
    "R": {"nota": 16, "letra": 17, "te": 19, "mes": 20, "ano": 21, "plantel": 22},
}
BLOQUES = [("L", 22, 31), ("R", 22, 31), ("L", 35, 44), ("R", 35, 44), ("L", 48, 57)]
for lado, r0, r1 in BLOQUES:
    col = COLS[lado]
    for r in range(r0, r1 + 1):
        for key in ("nota", "letra", "te", "mes", "ano", "plantel"):
            blank(ws.cell(row=r, column=col[key]))

# 4e. Componentes / Participación (derecha de 5º): valores del original.
for r in range(48, 62):
    for c in range(17, 23):  # Q..V (Nº, letra, T-E, mes, año, plantel)
        blank(ws.cell(row=r, column=c))

wb.save(DST)
print("OK ->", DST)
