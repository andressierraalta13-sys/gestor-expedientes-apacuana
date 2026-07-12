"""
Genera la plantilla Excel de Notas Certificadas 'FORMATO NC 10 MATERIAS.xlsx'
a partir del formato oficial 'FORMATO EN BLANCO IMPRESION (2).xlsx'.

La diferencia clave es que cada año (1º–5º) pasa a tener 10 filas de materias
(las 10 registradas en el sistema, en mayúsculas), en lugar de las 7/7/8/9/10
del formato anterior, para que las 10 asignaturas sean detectables y se rellenen
automáticamente. Se conserva todo el diseño (bordes, celdas combinadas,
encabezado con logos y firmas) clonando los estilos de las celdas originales.

Uso:  python scripts/build_nc_template.py   (desde la raíz del proyecto o cualquier
      sitio; las rutas se resuelven relativas a la raíz del repo).

Layout resultante (hoja 'Z Z FORMATO LIMPIO'):
  Filas 1-17 : encabezado institucional (SIN CAMBIOS)
  Bloque A   : título 18, subcab 19-20, materias 21-30  (1º izq / 2º der)
  Bloque B   : título 31, subcab 32-33, materias 34-43  (3º izq / 4º der)
  Bloque C   : título 44, subcab 45-46, materias 47-56  (5º izq ; Orientación/
               Participación der = filas 40-52 originales desplazadas +4)
  Fila 57    : separador
  Firmas     : 58-65 (filas 55-62 originales desplazadas +3; nombre A61, cédula A63)

Debe mantenerse en sincronía con apps/calificaciones/certificadas_generator.py
(constantes MATERIAS_SISTEMA, bloques_anos y coordenadas de firmas).
"""
import openpyxl, os, copy
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.styles import Font

AREA_FONT_SIZE = 8   # nombres de materia (más pequeño para que quepan sin recorte)
AREA_ROW_H = 31.5    # alto de fila de materia (fija hasta ~3 líneas a 8pt)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(BASE_DIR, "FORMATOS EXCEL", "FORMATO EN BLANCO IMPRESION (2).xlsx")
DST = os.path.join(BASE_DIR, "FORMATOS EXCEL", "FORMATO NC 10 MATERIAS.xlsx")
HOJA = "Z Z FORMATO LIMPIO"

SUBJECTS = [
    "LENGUA Y LITERATURA",
    "IDIOMAS",
    "MATEMÁTICA",
    "EDUCACIÓN FÍSICA",
    "BIOLOGÍA, AMBIENTE Y TECNOLOGÍA",
    "FÍSICA",
    "QUÍMICA",
    "GEOGRAFÍA, HISTORIA Y SOBERANÍA NACIONAL",
    "INNOVACIÓN TECNOLÓGICA Y PRODUCTIVA",
    "ORIENTACIÓN VOCACIONAL",
]

wb = openpyxl.load_workbook(SRC)
ws = wb[HOJA]

# ── 1. Snapshot de (valor, estilo, formato) y de los merges ORIGINALES ────────
snap = {}
for r in range(1, 63):
    for c in range(1, 25):
        cell = ws.cell(row=r, column=c)
        snap[(r, c)] = (cell.value, copy.copy(cell._style), cell.number_format)
orig_merges = [str(m) for m in ws.merged_cells.ranges]

def style_from(dst_r, dst_c, src_r, src_c, value="__keep__"):
    """Copia estilo+formato de la celda origen a la destino; fija valor si se indica."""
    sval, sstyle, snumfmt = snap[(src_r, src_c)]
    d = ws.cell(row=dst_r, column=dst_c)
    d._style = copy.copy(sstyle)
    d.number_format = snumfmt
    d.value = sval if value == "__keep__" else value

def shift(ref, drow):
    c1, r1, c2, r2 = range_boundaries(ref)
    return f"{get_column_letter(c1)}{r1+drow}:{get_column_letter(c2)}{r2+drow}"

# ── 2. Limpiar la región 18-70 (desunir merges, resetear valores/estilos) ─────
for mc in list(ws.merged_cells.ranges):
    if mc.min_row >= 18:
        ws.unmerge_cells(str(mc))
for r in range(18, 71):
    for c in range(1, 25):
        cell = ws.cell(row=r, column=c)
        cell.value = None
        cell.style = "Normal"

new_merges = []

# ── 3. Bloques de año (L y R) ─────────────────────────────────────────────────
# Fuente de estilos: título=fila18, subcab1=fila19, subcab2=fila20, materia=fila21
def build_year_L(title_row, title_text, subjects):
    t = title_row
    for c in range(1, 12):
        style_from(t, c, 18, c, value=title_text if c == 1 else None)
    new_merges.append(f"A{t}:K{t}")
    for c in range(1, 12):
        style_from(t + 1, c, 19, c)
        style_from(t + 2, c, 20, c)
    new_merges.extend([f"A{t+1}:C{t+2}", f"D{t+1}:F{t+1}", f"G{t+1}:G{t+2}",
                       f"H{t+1}:I{t+1}", f"J{t+1}:K{t+2}", f"E{t+2}:F{t+2}"])
    for i in range(10):
        r = t + 3 + i
        for c in range(1, 12):
            style_from(r, c, 21, c, value=subjects[i] if c == 1 else None)
        new_merges.extend([f"A{r}:C{r}", f"E{r}:F{r}", f"J{r}:K{r}"])

def build_year_R(title_row, title_text, subjects):
    t = title_row
    for c in range(12, 22):
        style_from(t, c, 18, c, value=title_text if c == 13 else None)
    new_merges.append(f"M{t}:U{t}")
    for c in range(12, 22):
        style_from(t + 1, c, 19, c)
        style_from(t + 2, c, 20, c)
    new_merges.extend([f"M{t+1}:N{t+2}", f"O{t+1}:Q{t+1}", f"R{t+1}:R{t+2}",
                       f"S{t+1}:T{t+1}", f"U{t+1}:U{t+2}", f"P{t+2}:Q{t+2}"])
    for i in range(10):
        r = t + 3 + i
        for c in range(12, 22):
            style_from(r, c, 21, c, value=subjects[i] if c == 13 else None)
        new_merges.extend([f"M{r}:N{r}", f"P{r}:Q{r}"])

build_year_L(18, "PRIMER AÑO", SUBJECTS)
build_year_L(31, "TERCER AÑO", SUBJECTS)
build_year_L(44, "QUINTO AÑO", SUBJECTS)
build_year_R(18, "SEGUNDO AÑO", SUBJECTS)
build_year_R(31, "CUARTO AÑO", SUBJECTS)

# ── 4. Bloque C derecho: Orientación + Participación (filas 40-52 orig, +4) ────
for src_r in range(40, 53):
    dst_r = src_r + 4
    for c in range(12, 25):
        style_from(dst_r, c, src_r, c)  # conserva valores y estilos originales
for m in orig_merges:
    c1, r1, c2, r2 = range_boundaries(m)
    if 40 <= r1 <= 52 and c1 >= 13:
        new_merges.append(shift(m, 4))

# ── 5. Separador (fila 57) ────────────────────────────────────────────────────
for c in range(1, 25):
    style_from(57, c, 54, c, value=None)
new_merges.append("A57:U57")

# ── 6. Firmas: filas 55-62 originales desplazadas +3 → 58-65 ─────────────────
for src_r in range(55, 63):
    dst_r = src_r + 3
    for c in range(1, 25):
        style_from(dst_r, c, src_r, c)
for m in orig_merges:
    c1, r1, c2, r2 = range_boundaries(m)
    if 55 <= r1 <= 62:
        new_merges.append(shift(m, 3))

# ── 7. Aplicar merges (sin duplicados) ────────────────────────────────────────
for m in sorted(set(new_merges)):
    ws.merge_cells(m)

# ── 8. Alturas de fila, anchos y page setup ───────────────────────────────────
def set_h(r, h):
    ws.row_dimensions[r].height = h
for t in (18, 31, 44):
    set_h(t, 12.75)
    set_h(t + 1, 18.0)
    set_h(t + 2, 14.25)
    for i in range(10):
        set_h(t + 3 + i, AREA_ROW_H)
for r in range(58, 66):
    if ws.row_dimensions[r].height is None:
        ws.row_dimensions[r].height = 14.0

# Ensanchar un poco la columna de nombre de materia (C y N).
ws.column_dimensions['C'].width = 9.5
ws.column_dimensions['N'].width = 11.5

# Page setup: ajustar a 1 página de ancho al imprimir/exportar (Oficio vertical).
ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.print_area = 'A1:U65'

# Reducir el tamaño de fuente del NOMBRE de materia para que los nombres largos
# (BIOLOGÍA, AMBIENTE Y TECNOLOGÍA / GEOGRAFÍA... / INNOVACIÓN...) quepan completos.
for t in (18, 31, 44):
    for i in range(10):
        r = t + 3 + i
        for col in (1, 13):  # A (izq) y M (der)
            cell = ws.cell(row=r, column=col)
            f = cell.font
            cell.font = Font(name=f.name, size=AREA_FONT_SIZE, bold=f.bold,
                             italic=f.italic, color=f.color)

wb.save(DST)
print("OK ->", DST)
